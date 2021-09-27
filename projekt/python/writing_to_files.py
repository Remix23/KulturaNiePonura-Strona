from openpyxl import Workbook
from string import ascii_uppercase

def log_persons (persons : list, selector : dict,  workbook = None, title = 'Osoby'):
      
    if isinstance(workbook, Workbook):
        wb = workbook
    else:
        wb = Workbook()

    ws = wb.create_sheet(title = title)

    to_log = [key.replace("-", "_") for key in selector if selector[key]]

    ws['A1'] = 'ID'

    headers = ["imię", "nazwisko", "pseudomin", "miejsce zamieszkania", "adres e-mail",
        "numer telefonu", "opis", "opis twórczości", "link do prac", "link do konkatu",
        "notatki", "instytucje", "rodzaje sztuki", "proponowany rodzaj promocji", 
        "zna festiwal z", "rok urodzenia", "data zerejestrowania"
    ]

    for i, k in enumerate(to_log):
        ws[f"{ascii_uppercase[i + 1]}1"] = s.headers[k]

    for row, person in enumerate(persons):
        row_num = row + 2
        _ = ws.cell(row = row_num, column = 1, value = person['id'])
        
        for i, key in enumerate(to_log):
            if key in ['instytutions', 'art_kinds', 'promotion_types', 'know_from_types']:
                _ = ws.cell(row = row_num, column = i + 2, value = ", ".join(person[key]))
            else:
                _ = ws.cell(row = row_num, column = i + 2, value = person[key])

def log_teams (teams : list, selector : dict, workbook = None, title = 'Zespoły'):
    
    if isinstance(workbook, Workbook):
        wb = workbook
    else: 
        wb = Workbook()

    ws = wb.create_sheet(title = title)

    to_log = [key.replace("-", "_") for key in selector if selector[key]]

    ws['A1'] = 'ID'

    for i, k in enumerate(to_log):
        ws[f"{ascii_uppercase[i + 1]}1"] = s.headers[k]

    for row, team in enumerate(teams):
        row_num = row + 2
        _ = ws.cell(row = row_num, column = 1, value = team['id'])

        for i, key in enumerate(to_log):
            if key in ['instytutions', 'art_kinds', 'promotion_types', 'know_from_types', 'born_dates', 'living_places']:
                _ = ws.cell(row = row_num, column = i + 2, value = ", ".join(team[key]))
            else:
                _ = ws.cell(row = row_num, column = i + 2, value = team[key])

def calculate_stats (persons : list, teams : list):

    instytutions = {}
    art_kinds = {}    
    promotion_types = {}    
    know_from_types = {}
    for line in persons + teams:
        for instytution in line['instytutions']:
            if instytution in instytutions:
                instytutions[instytution] += 1
            else:
                instytutions[instytution] = 1

        for art_kind in line['art_kinds']:
            if art_kind in art_kinds:
                art_kinds[art_kind] += 1
            else:
                art_kinds[art_kind] = 1

        for promo_type in line['promotion_types']:
            if promo_type in promotion_types:
                promotion_types[promo_type] += 1
            else:
                promotion_types[promo_type] = 1

        for know_from_type in line['know_from_types']:
            if know_from_type in know_from_types:
                know_from_types[know_from_type] += 1
            else:
                know_from_types[know_from_type] = 1

    living_places = {}
    for team in teams:
        for place in team['living_places']:
            if place in living_places:
                living_places[place] += 1
            else:
                living_places[place] = 1
    for person in persons:
        if person['living_place'] in living_places:
            living_places[person['living_place']] += 1
        else:
            living_places[person['living_place']] = 1

    born_dates = {}
    for team in teams:
        for date in team['born_dates']:
            if date in born_dates:
                born_dates[date] += 1
            else:
                born_dates[date] = 1
    for person in persons:
        if person['year_born'] in born_dates:
            born_dates[person['year_born']] += 1
        else:
            born_dates[person['year_born']] = 1
    
    return instytutions, art_kinds, promotion_types, know_from_types, born_dates, living_places

def log_stats (instytutions, art_kinds, promotion_types, know_from_types, born_dates, living_places, workbook = None, title = 'Ogólne', first = False):
    
    if isinstance(workbook, Workbook):
        wb = workbook
    else: 
        wb = Workbook()
    if first:
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title = title)

    ws['A1'] = 'Instytucje'
    ws['B1'] = 'Ilośc wystąpień'
    ws['C1'] = 'Rodzaje sztuki'
    ws['D1'] = 'Ilośc wystąpień'
    ws['E1'] = 'Sposoby promocji'
    ws['F1'] = 'Ilośc wystąpień'
    ws['G1'] = 'Skąd wiedzą o festiwalu'
    ws['H1'] = 'Ilośc wystąpień'
    ws['I1'] = 'Miejsca zamieszkania'
    ws['J1'] = 'Ilośc wystąpień'
    ws['K1'] = 'Daty urodzenia'
    ws['L1'] = 'Ilośc wystąpień'

    for row, instytution in enumerate(instytutions):
        _ = ws.cell(row = row + 2, column = 1, value = instytution)
        _ = ws.cell(row = row + 2, column = 2, value = instytutions[instytution])

    for row, art_kind in enumerate(art_kinds):
        _ = ws.cell(row = row + 2, column = 3, value = art_kind)
        _ = ws.cell(row = row + 2, column = 4, value = art_kinds[art_kind])

    for row, promotion_type in enumerate(promotion_types):
        _ = ws.cell(row = row + 2, column = 5, value = promotion_type)
        _ = ws.cell(row = row + 2, column = 6, value = promotion_types[promotion_type])

    for row, know_from_type in enumerate(know_from_types):
        _ = ws.cell(row = row + 2, column = 7, value = know_from_type)
        _ = ws.cell(row = row + 2, column = 8, value = know_from_types[know_from_type])

    for row, living_place in enumerate(living_places):
        _ = ws.cell(row = row + 2, column = 9, value = living_place)
        _ = ws.cell(row = row + 2, column = 10, value = living_places[living_place])

    for row, born_date in enumerate(born_dates):
        _ = ws.cell(row = row + 2, column = 11, value = born_date)
        _ = ws.cell(row = row + 2, column = 12, value = born_dates[born_date])

def log_all_to_excel (persons, teams, filename):

    wb = Workbook()

    instytutions, art_kinds, promotion_types, know_from_types, born_dates, living_places = calculate_stats(persons, teams)

    log_stats (instytutions, art_kinds, promotion_types, know_from_types, born_dates, living_places, wb, first=True)

    log_persons(persons, wb)
    log_teams(teams, wb)

    wb.save(filename = filename)

def log_selected_to_excel (persons, teams, filename, selector):
    
    wb = Workbook()

    if selector["stats"]:
        instytutions, art_kinds, promotion_types, know_from_types, born_dates, living_places = calculate_stats(persons, teams)

        log_stats (instytutions, art_kinds, promotion_types, know_from_types, born_dates, living_places, wb, first=True)
    del selector['stats']
    persons_selector = {key:value for (key, value) in selector.items() if not key.startswith('team')}
    teams_selector = {key[5:]:value for (key, value) in selector.items() if key.startswith('team')}
    
    log_persons(persons = persons, selector = persons_selector, workbook = wb)
    log_teams(teams = teams, selector = teams_selector, workbook = wb)
    wb.save(filename = filename)

def log_all_to_html (persons, teams, filename):
    pass

def log_single_to_excel(data : dict, filename, title = 'Dane szczegółowe'):

    wb = Workbook()

    ws = wb.active
    ws.title = title

    row = 1

    for key, value in data.items():
        if type(value) != list:
            _ = ws.cell(column = 1, row = row, value = key)
            _ = ws.cell(column = 2, row = row, value = value)
        
        else:
            _ = ws.cell(column = 1, row = row, value = key)
            _ = ws.cell(column = 2, row = row, value = ', '.join(value))

        row += 1
    
    wb.save(filename = filename)
